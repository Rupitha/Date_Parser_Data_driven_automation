import openpyxl
import pytest


@pytest.fixture()
class Dateparser:

    @staticmethod
    def getdata():
        Dict = {}

        book = openpyxl.load_workbook("//home//rupz//Documents//inputs.xlsx")
        sheet = book.active
        cell = sheet.cell(row=1,column=2)
        print(cell.value)
        for i in range(2,sheet.max_row+1):
            print(sheet.cell(row=i,column=2).value)
            #print(sheet.cell(row=i,column=sheet.max_column).value)
            Dict[sheet.cell(row=1,column=2).value]=sheet.cell(row=i,column=2).value
        return Dict



   



