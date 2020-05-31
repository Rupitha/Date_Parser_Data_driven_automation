import openpyxl

class Dateparser:
    @staticmethod
    def getdata():


        book = openpyxl.load_workbook("TestData/inputs.xlsx")
        sheet = book.active
        cell = sheet.cell(row=1,column=2)
        print(cell.value)
        ret = []
        for i in range(2, sheet.max_row + 1):
            Dict = {}
            name = sheet.cell(row=i, column=3).value
            input = sheet.cell(row=i, column=2).value
            expected = sheet.cell(row=i, column=6).value
            if(input is not None):
                Dict[sheet.cell(row=1, column=3).value.strip()]=name
                Dict[sheet.cell(row=1, column=2).value.strip()]=input
                Dict[sheet.cell(row=1,column=6).value.strip()]=expected
                ret.append(Dict)
            #return [Dict]
        return ret
        # return [
        #     {'Test Case Description': 'Test to validate MM/DD/YYYY', 'Test Inputs': '18/12/1994', ' Expected Result': 'Return result Day Sun Dec 18 1994 00:00:00 GMT+0000'},
        #     {'Test Case Description': 'Test to validate MM/DD/YYYY', 'Test Inputs': '18/12/1994', ' Expected Result': 'Return result Day Sun Dec 18 1994 00:00:00 GMT+0000'}
        #         ]






